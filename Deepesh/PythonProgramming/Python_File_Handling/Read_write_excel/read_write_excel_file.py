# install openpyxl package
# pip install openpyxl

import openpyxl

def read_excel_file(filepath, sheet_name, cell_name):
    wb = openpyxl.load_workbook(filepath)
    #print(wb.__dict__)
    sheet = wb[sheet_name]
    #print(sheet.__dict__)
    cell = sheet[cell_name]
    print(cell.value)


read_excel_file("test_data.xlsx", "Sheet1", "A1") # Pune


# for i in range(1, 5):
#     read_excel_file("test_data.xlsx", "Sheet1", f"A{i}")
"""
Pune
Mumbai
Bangalore
Chennai
"""

# Get all the sheet values with sheet number
# for i in range(1, 4):
#     for j in range(1, 5):
#         read_excel_file("test_data.xlsx", f"Sheet{i}", f"A{j}")
#     print("_"*40)



# Get all the sheet values with sheet name
# for sheet in ['Sheet1', 'Sheet2', 'Sheet3']:
#     for j in range(1, 5):
#         read_excel_file("test_data.xlsx", sheet, f"A{j}")
#     print("_"*40)
#


####################################
# Get all values in one single colum

def get_colum_values_excel_file(filepath, sheet_name, colum_name):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[sheet_name]
    max_colns = sheet.max_column
    max_rows = sheet.max_row
    print("max colums :", max_colns)
    print("max rows :", max_rows)
    for i in range(1, max_rows+1):
        cell = sheet[f"{colum_name}{i}"]
        print(cell.value)


#get_colum_values_excel_file("test_data.xlsx", "Sheet1", "A")
#get_colum_values_excel_file("test_data.xlsx", "Sheet3", "D")


# Write content to excel sheet
def write_content_to_excel_sheet(filepath, sheet_name, cell_name, new_value):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    cell.value = new_value
    wb.save(filepath)


#write_content_to_excel_sheet("test_data.xlsx", "Sheet1", "A12", "Good Evening")



def write_content_next_available_cell(filepath, sheet_name, colum_name, new_value):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[sheet_name]
    max_rows = sheet.max_row
    cell = sheet[f"{colum_name}{max_rows+1}"]
    cell.value = new_value
    wb.save(filepath)

write_content_next_available_cell("test_data.xlsx", "Sheet1", "B", "Python Programming")
