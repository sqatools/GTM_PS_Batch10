import openpyxl


def read_excel_file(filepath, sheet_name, cell_name):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    print(cell.value)


read_excel_file("test_data.xlsx", "Sheet1", "A1")
# Pune


print("-"*50)
for i in range(1, 6):
    read_excel_file("test_data.xlsx", "Sheet1", f"A{i}")
# Pune
# Mumbai
# Bangalore
# Delhi
# Panjab


print("-"*50)
for i in range(1, 3):
    for j in range(1, 5):
        read_excel_file("test_data.xlsx", f"Sheet{i}", f"A{j}")

"""
Pune
Mumbai
Bangalore
Delhi
Apple
Mango
Lemon
Orange
"""


print("-"*50)
for sheet in ['Sheet1', 'Sheet2']:
    for j in range(1, 5):
        read_excel_file("test_data.xlsx", sheet, f"A{j}")
# Pune
# Mumbai
# Bangalore
# Delhi
# Apple
# Mango
# Lemon
# Orange


print("-"*50)


def get_colum_values_excel_file(filepath, sheet_name, colum_name):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[sheet_name]
    max_colns = sheet.max_column
    max_rows = sheet.max_row
    print("max columns :", max_colns)
    print("max rows :", max_rows)
    for i in range(1, max_rows + 1):
        cell = sheet[f"{colum_name}{i}"]
        print(cell.value)


# get_colum_values_excel_file("test_data.xlsx", "Sheet1", "A")
# get_colum_values_excel_file("test_data.xlsx", "Sheet3", "D")


def write_content_to_excel_sheet(filepath, sheet_name, cell_name, new_value):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    cell.value = new_value
    wb.save(filepath)


# write_content_to_excel_sheet("test_data.xlsx", "Sheet1", "A11", "Good Morning")


def write_content_next_available_cell(filepath, sheet_name, colum_name, new_value):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[sheet_name]
    max_rows = sheet.max_row
    cell = sheet[f"{colum_name}{max_rows+1}"]
    cell.value = new_value
    wb.save(filepath)

write_content_next_available_cell("test_data.xlsx", "Sheet1", "B", "Python Programming")
