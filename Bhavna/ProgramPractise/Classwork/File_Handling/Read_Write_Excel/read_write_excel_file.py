# 02/07/2025 Session

# install openpyxl package
# pip install openpyxl

import openpyxl

def read_excel_file(filepath,sheet_name,cell_name):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    print(cell.value)

# read_excel_file("test_data.xlsx","Sheet1","A2") # Abu Dhabi

# for i in range(1,4):
#     read_excel_file("test_data.xlsx","Sheet1",f"A{i}")
'''
Dubai
Abu Dhabi
Ras al khaimah
'''

# Get all the sheet values with sheet number
'''
for i in range(1,4):
    for j in range(1,5):
        read_excel_file("test_data.xlsx",f"Sheet{i}",f"A{j}")
    print('_'*70)
'''

# Get all the sheet values with sheet name
for sheet in ['Sheet1','Sheet2','Sheet3']:
    for j in range(1,5):
        read_excel_file("test_data.xlsx",sheet,f"a{j}")
    print('_'*70)

####################################
# Get all values in one single column

def get_column_value(filepath,sheet_name,column_name):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[sheet_name]
    max_column = sheet.max_column
    max_rows = sheet.max_row
    print("max column:",max_column)
    print("max rows:",max_rows)
    for i in range(1, max_rows+1):
        cell = sheet[f"{column_name}{i}"]
        print(cell.value)

# get_column_value("test_data.xlsx","Sheet1","A")
# get_column_value("test_data.xlsx","Sheet3","B")

# Write content to excel sheet

def write_content_to_excel(filepath,sheet_name,cell_name,new_value):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    cell.value = new_value
    wb.save(filepath)

# write_content_to_excel("test_data.xlsx","Sheet2","B1","Turquoise")


def write_content_next_available_cell(filepath,sheet_name,column_name,new_value):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[sheet_name]
    max_rows = sheet.max_row
    cell = sheet[f"{column_name}{max_rows+1}"]
    cell.value = new_value
    wb.save(filepath)

write_content_next_available_cell("test_data.xlsx","Sheet2","C","Hello")
